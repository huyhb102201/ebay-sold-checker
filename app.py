from flask import Flask, request, render_template, jsonify
import undetected_chromedriver as uc
from openpyxl import Workbook, load_workbook
import re, time, os

app = Flask(__name__)

def get_sold_quantity(url):
    options = uc.ChromeOptions()
    options.headless = True
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--log-level=3")
    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.managed_default_content_settings.stylesheets": 2,
    }
    options.add_experimental_option("prefs", prefs)

    driver = uc.Chrome(options=options)
    driver.get(url)
    time.sleep(3)
    html = driver.page_source
    driver.quit()

    match = re.search(r"(\d+)\s+sold", html)
    return match.group(1) if match else "Không tìm thấy"

def write_to_excel(url, sold):
    filename = "data.xlsx"
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "eBay Data"
        ws.append(["URL", "Sold Quantity"])
    else:
        wb = load_workbook(filename)
        ws = wb.active
    ws.append([url, sold])
    wb.save(filename)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/check", methods=["POST"])
def check():
    data = request.get_json()
    url = data.get("ebay_url")
    try:
        sold = get_sold_quantity(url)
        write_to_excel(url, sold)
        return jsonify({"sold": sold})
    except Exception as e:
        return jsonify({"sold": f"Lỗi: {str(e)}"})

if __name__ == "__main__":
    app.run(debug=True)
